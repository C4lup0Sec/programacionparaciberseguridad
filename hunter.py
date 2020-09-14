""" CARLOS Mz H
"""
import sys
from pyhunter import PyHunter
from openpyxl import Workbook

# Sustituye la siguiente API key por la tuya
HUNTER = PyHunter('hgghbk58e1gdvfvfvfdvf98254hjgh7')


def busqueda(organizacion):
    """ Cantidad de resultados esperados de la búsqueda 1,
    El límite MENSUAL de Hunter es 50, cuidado! """
    search = 1
    resultado = HUNTER. domain_search(company=organizacion,
                                      limit=search, emails_type='personal')
    return resultado


def guardar_informacion(_datos_encontrados, organizacion):
    """ metodo encargado de guardar los resultados en un archivo de excel """
    libro = Workbook()
    hoja = libro.create_sheet(organizacion)
    libro.save("Hunter" + organizacion + ".xlsx")
    hoja.cell(1, 1, "Dominio")
    hoja.cell(1, 2, DATOS_ENCONTRADOS['domain'])
    hoja.cell(2, 1, "Organización")
    hoja.cell(2, 2, DATOS_ENCONTRADOS['organization'])
    hoja.cell(3, 1, "Correo")
    hoja.cell(3, 2, DATOS_ENCONTRADOS['emails'][0]['value'])
    hoja.cell(4, 1, "Nombre(s)")
    hoja.cell(4, 2, DATOS_ENCONTRADOS['emails'][0]['first_name'])
    hoja.cell(5, 1, "Apellidos")
    hoja.cell(5, 2, DATOS_ENCONTRADOS['emails'][0]['last_name'])
    libro.save("Hunter" + organizacion + ".xlsx")

print("Script para buscar información")
ORGA = input("Organización a investigar: ")
DATOS_ENCONTRADOS = busqueda(ORGA)
if DATOS_ENCONTRADOS is None:
    sys.exit()
else:
    print(DATOS_ENCONTRADOS)
    print(type(DATOS_ENCONTRADOS))
    guardar_informacion(DATOS_ENCONTRADOS, ORGA)
